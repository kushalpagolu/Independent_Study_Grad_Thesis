import os
import time
import pandas as pd
import logging
from kalman_filter import KalmanFilter
import numpy as np
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def save_data_continuously(data_store, stop_saving_thread, data_lock, raw_path, processed_path):
    """
    Continuously saves EEG + gyro data from the data_store to Excel files every 30 seconds.

    - raw_path: Excel file for raw values (14 EEG channels + 2 gyro)
    - processed_path: Excel file for data with computed volts, filtering, smoothing

    Both files are created once with headers, then each batch is appended.
    """
    # Ensure output directories exist
    os.makedirs(os.path.dirname(raw_path), exist_ok=True)
    os.makedirs(os.path.dirname(processed_path), exist_ok=True)

    # Channel definitions
    eeg_channels  = ["AF3","F7","F3","FC5","T7","P7","O1","O2","P8","T8","FC6","F4","F8","AF4"]
    gyro_channels = ["gyro_x","gyro_y"]
    raw_cols      = eeg_channels + gyro_channels

    # Create raw file with header if missing
    if not os.path.exists(raw_path):
        pd.DataFrame(columns=raw_cols).to_excel(raw_path, index=False)
        #logger.info(f"[Data Saver] Created raw data file with headers: {raw_path}")

    logging.info("[Data Saver] save_data_continuously thread started.")

    try:
        while not stop_saving_thread.is_set():
            # Grab and clear the shared store
            with data_lock:
                local_data = data_store.copy()
                data_store.clear()

            if local_data:
                #logger.info(f"[Data Saver] Collected {len(local_data)} rows of data.")

                # Build DataFrame for raw data
                df_raw = pd.DataFrame(local_data, columns=raw_cols)

                # Copy for processing
                df = df_raw.copy()

                # Compute EEG channel volts
                for ch in eeg_channels:
                    df[f"{ch}_volts"] = df[ch] * 0.51

                # Kalman filter + integrate gyro
                kf_x = KalmanFilter()
                kf_y = KalmanFilter()
                df["gyro_x_deg_s"] = df["gyro_x"].apply(lambda x: kf_x.update(x) if pd.notna(x) else np.nan)
                df["gyro_y_deg_s"] = df["gyro_y"].apply(lambda y: kf_y.update(y) if pd.notna(y) else np.nan)
                df["head_roll_deg"]  = df["gyro_x_deg_s"].cumsum().fillna(0) * (1/128)
                df["head_pitch_deg"] = df["gyro_y_deg_s"].cumsum().fillna(0) * (1/128)

                # Median subtraction
                med = df[eeg_channels].median(axis=1)
                for ch in eeg_channels:
                    df[f"{ch}_med_subtracted"] = df[ch] - med

                # Clip & smooth EEG
                for i in range(1, len(df)):
                    delta = df.loc[i, eeg_channels] - df.loc[i-1, eeg_channels]
                    delta = delta.clip(-15, 15)
                    df.loc[i, eeg_channels] = df.loc[i-1, eeg_channels] + delta

                # Create processed file with headers if missing
                if not os.path.exists(processed_path):
                    pd.DataFrame(columns=df.columns.tolist()).to_excel(processed_path, index=False)
                    #logger.info(f"[Data Saver] Created processed data file with headers: {processed_path}")

                # Append raw rows
                wb_raw = load_workbook(raw_path)
                ws_raw = wb_raw.active
                for row in df_raw.itertuples(index=False, name=None):
                    ws_raw.append(row)
                wb_raw.save(raw_path)

                # Append processed rows
                wb_proc = load_workbook(processed_path)
                ws_proc = wb_proc.active
                for row in df.itertuples(index=False, name=None):
                    ws_proc.append(row)
                wb_proc.save(processed_path)

                #logger.info(f"[Data Saver] Appended to {raw_path} and {processed_path}")
            else:
                logger.debug("[Data Saver] No new data this cycle.")

            time.sleep(30)
    except Exception as e:
        logger.error(f"[Data Saver] Error in save_data_continuously: {e}")
    finally:
        logger.info("[Data Saver] save_data_continuously thread stopping.")
